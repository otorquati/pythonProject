import openpyxl

# Abre o arquivo e cola em uma variável workbook
inv_file = openpyxl.load_workbook("./Automation/inventory.xlsx")

#Coloca uma planilha específica em variável
product_list = inv_file['Sheet1']

# Exercício 1: Listar cada comapnhia e seus recpectivos produtos
# Cria um dicionário vazio de produtos por fornecedor
products_per_supplier = {}
total_value_per_supplier = {}
products_under_10inv = {}

# Cria um loop para a faixa "range" de 2 a 75 (max_row)
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)
    # Calcula número de produtos por fornecedor
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier[supplier_name]
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        # print("Adicionando Fornecedor")
        products_per_supplier[supplier_name] = 1

# Exercício 2: Calcular quantidade de produtos por fornecedor
    if supplier_name in total_value_per_supplier:
       current_total_value = total_value_per_supplier.get(supplier_name) 
       total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

# Exercício 3: produtos com qutd menro que 10
    if inventory < 10:
        products_under_10inv[product_number]=int(inventory)

# Exercício 4: Escrever resultados na planilha
    # Adicionando valores 
    inventory_price.value = price * inventory


# Resultado Exercício 1
print(products_per_supplier)
# Resultado Exercício 2
print(total_value_per_supplier)
# Resultado Exercício 3
print(products_under_10inv)
# Resultado Exercício 4
inv_file.save("./Automation/inventory_updated.xlsx")